from openpyxl import Workbook, load_workbook
import os
from io import BytesIO


class BuildData(object):
    """Organizes data from AM build spreadsheet templates into object with dicts that can be imported into MongoDB.

    The BuildData object stores spreadsheet data as entries (row-wise) and trees (hierarchical build structure).
    Inheritance is based on the UID naming structure, e.g. CMU01.01 = Part 01 within Build CMU01. Entries and trees are
    stored as dicts that can be imported into MongoDB or processed externally. The input is a file-like object, e.g. f.read().

    Args:
        file_obj: A file-like object containing the AM build template spreadsheet (.xlsx) template, e.g. f.read().
        filename: The name of the spreadsheet.
        project_id: The project id that this build will be associated with.

    Returns:
        An object storing build data as entries and a tree.
    """
    def __init__(self, file_obj, filename, project_id):
        (file, ext) = os.path.splitext(os.path.basename(filename))
        # wb = load_workbook(filename, data_only=True)
        wb = load_workbook(BytesIO(file_obj.read()), data_only=True)
        self.data = dict()
        self.entries = []
        self.name = file
        self.data.update({'project_id': project_id})
        self.data.update({'_id': file})
        self.data.update({'UID': file})
        self.data.update({'files': []})
        for i in range(len(wb.sheetnames)):
            ws = wb.worksheets[i]
            sheet_data = []
            column_list = []
            for col in ws.iter_cols(min_row=2, max_row=2):
                for cell in col:
                    column_list.append(cell.value)

            for row in ws.iter_rows(min_row=5):
                row_dict = dict()
                row_dict.update({'files': []})
                value_list = []
                for cell in row:
                    value_list.append(cell.value)
                for j in range(len(column_list)):
                    key = column_list[j].replace('.', '')
                    row_dict.update({key: value_list[j]})
                sheet_data.append(row_dict)
                entry_dict = {'_id': row_dict['UID'], 'sheet': ws.title, 'files': [], 'project_id': project_id}
                entry_dict.update(row_dict)
                self.entries.append(entry_dict)
                try:
                    for k in range(len(self.data['Part'])):
                        if 'UID' in self.data['Part'][k] and row_dict['FID'] == self.data['Part'][k]['UID']:
                            self.data['Part'][k].update({ws.title: row_dict})
                except KeyError:
                    pass
            self.data.update({ws.title: sheet_data})
        self.entries.append(self.data)

